import os
import requests
from datetime import datetime, timezone
from tqdm import tqdm
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import re
from dotenv import load_dotenv

load_dotenv()

# ===================== CONFIGURAÇÕES =====================
API_KEY = os.getenv("API_KEY")
TOKEN = os.getenv("TOKEN")
LIST_ID = os.getenv("LIST_ID")

ONEDRIVE_PATH = r"C:\Users\Camim\OneDrive\NotaAtualizacao"
MODELO_PATH = os.path.join(ONEDRIVE_PATH, "MODELO.xlsx")

# Cia subpasta para as planilhas
PLANILHAS_PATH = os.path.join(ONEDRIVE_PATH, "Planilhas")
os.makedirs(PLANILHAS_PATH, exist_ok=True)

DATA_HOJE = datetime.now().strftime("%d%m%Y-%H%M")
ARQUIVO_EXCEL = os.path.join(PLANILHAS_PATH, f"relatorio{DATA_HOJE}.xlsx")

PATTERN_CHAMADO = "https://servicedesk.clinicacamim.com.br/admin/admin_ticket.php?track="

# ===================== FUNÇÕES DE TRATAMENTO DE TEXTO =====================

def coluna_solicitante(descricao):
    """Extrai o solicitante/requerente da descrição"""
    import re
    if not isinstance(descricao, str) or not descricao.strip():
        return ""
    match = re.search(r"(Solicitante|Requerente|Nome)[:\-]\s*(.+)", descricao, re.IGNORECASE)
    if match:
        return match.group(2).replace("*", "").strip()
    return ""

def coluna_pedido(texto):
    """Limpa a descrição para a coluna Pedido"""
    if not isinstance(texto, str) or not texto.strip():
        return ""
    
    texto = re.sub(r"(Solicitante|Requerente|Nome)[:\-]\s*.+", "", texto, flags=re.IGNORECASE)
    texto = re.sub(r".*?(Problema|Incidente|Requisição|Pedido)[:\-]?\s*", "", texto, flags=re.IGNORECASE)
    texto = re.sub(r"[*_`~]+", "", texto)
    texto = re.sub(r"https?://\S+", "", texto)
    texto = re.sub(r'!\[.*?\]\(.*?\)', '', texto)
    texto = re.sub(r"\s+", " ", texto).strip()

    if texto and not texto[0].isupper():
        texto = texto[0].upper() + texto[1:]
    if texto and not texto.endswith((".", "!", "?")):
        texto += "."
    return texto

def coluna_observacao(card_id, data_done=None):
    """Retorna comentário e imagem para a coluna Observação"""
    autores_permitidos = {"victoriautrini", "Marcelle Santos"}
    url = f"https://api.trello.com/1/cards/{card_id}/actions"
    params = {"key": API_KEY, "token": TOKEN, "filter": "commentCard"}
    r = requests.get(url, params=params)
    if r.status_code != 200:
        return "", ""
    comentarios = r.json()
    if not comentarios:
        return "", ""

    # Converter data do DONE para datetime
    data_done_dt = None
    if data_done:
        try:
            data_done_dt = datetime.strptime(data_done, "%Y-%m-%dT%H:%M:%S.%fZ").replace(tzinfo=None)
        except Exception:
            try:
                data_done_dt = datetime.strptime(data_done, "%Y-%m-%dT%H:%M:%S.%f%z").astimezone().replace(tzinfo=None)
            except Exception:
                pass

    for c in comentarios:
        autor = c.get("memberCreator", {}).get("fullName", "")
        texto = c.get("data", {}).get("text", "")
        data_coment_str = c.get("date")
        try:
            data_coment = datetime.strptime(data_coment_str, "%Y-%m-%dT%H:%M:%S.%fZ").replace(tzinfo=None)
        except Exception:
            continue
        if autor in autores_permitidos and (not data_done_dt or data_coment > data_done_dt):
            link_imagem = ""
            match = re.search(r'!\[.*?\]\((.*?)\)', texto)
            if match:
                link_imagem = match.group(1)
            if not link_imagem:
                match = re.search(r'(https?://\S+\.(?:jpg|jpeg|png|gif))', texto)
                if match:
                    link_imagem = match.group(1)
            texto_limpo = re.sub(r'!\[.*?\]\(.*?\)', '', texto).strip()
            return texto_limpo, link_imagem
    return "", ""

# ===================== FUNÇÕES TRELLO =====================
def buscar_cards():
    url = f"https://api.trello.com/1/lists/{LIST_ID}/cards"
    params = {"key": API_KEY, "token": TOKEN}
    r = requests.get(url, params=params)
    r.raise_for_status()
    return r.json()

def buscar_anexo_filtrado(card_id):
    url = f"https://api.trello.com/1/cards/{card_id}/attachments"
    params = {"key": API_KEY, "token": TOKEN}
    r = requests.get(url, params=params)
    if r.status_code == 200 and r.json():
        for a in r.json():
            link = a.get("url")
            if link and link.startswith(PATTERN_CHAMADO):
                return link
    return None

def buscar_data_done(card_id):
    url = f"https://api.trello.com/1/cards/{card_id}/actions"
    params = {"key": API_KEY, "token": TOKEN, "filter": "updateCard:idList"}
    r = requests.get(url, params=params)
    if r.status_code != 200:
        return None

    done_dates = []
    for ac in r.json():
        data = ac.get("data", {})
        list_after = data.get("listAfter", {}).get("name", ""). strip().lower()
        if "done" in list_after:
            done_dates.append(ac.get("date"))

    if not done_dates:
        return None
    return max(done_dates)

def formatar_data_sem_hora(data_str):
    if not data_str:
        return ""
    try:
        dt = datetime.strptime(data_str, "%Y-%m-%dT%H:%M:%S.%fZ")
        return dt.strftime("%d/%m/%Y")
    except Exception:
        return data_str

# ===================== GERAR PLANILHA =====================
def ajustar_layout(ws):
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for col in ws.columns:
        max_length = 0
        coluna = col[0].column
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[get_column_letter(coluna)].width = max_length + 5

def gerar_planilha():
    cards = buscar_cards()
    # 🛑 ADICIONE ESTA LINHA DE DIAGNÓSTICO
    print(f"\nTotal de cards encontrados na lista (LIST_ID): {len(cards)}")
    if len(cards) <= 1:
        print("AVISO: A lista do Trello pode estar vazia ou o ID da lista está incorreto.")
    # FIM DO DIAGNÓSTICO
    
    rows = []
    for card in tqdm(cards[1:], desc="Processando cards"):
        titulo = card.get("name")
        link_card = card.get("shortUrl")
        anex = buscar_anexo_filtrado(card.get("id"))
        data_done_raw = buscar_data_done(card.get("id"))
        data_done = formatar_data_sem_hora(data_done_raw)

        # Etiquetas (labels)
        labels = card.get("labels", [])
        nomes_labels = [lbl.get("name", "").strip() for lbl in labels if lbl.get("name")]
        tipo = ", ".join(nomes_labels) if nomes_labels else "Sem etiqueta"

        # Extrair protocolo do título (padrão: letras/números, possivelmente entre colchetes)
        protocolo_match = re.search(r'(\[?[A-Z0-9]{2,}-[A-Z0-9]{2,}-[A-Z0-9]{2,}\]?)$', titulo)
        protocolo = protocolo_match.group(1) if protocolo_match else ""
        titulo_limpo = re.sub(r'\s*[-–—]?\s*' + re.escape(protocolo) + r'$', '', titulo).strip() if protocolo else titulo.strip()

        # Descrição limpar, sem formatação do Trello
        descricao = card.get("desc", "")
        descricao_limpa = coluna_pedido(descricao)

        # Comentário e imagem
        comentario, imagem_comentario = coluna_observacao(card.get("id"), data_done_raw)

        # 🛑 ADICIONE ESTA LINHA DE DIAGNÓSTICO DENTRO DO LOOP
        if not nomes_labels:
            print(f"Card '{titulo}' não tem etiquetas. 'Tipo' será 'Sem etiqueta'.")
        # FIM DO DIAGNÓSTICO

        # Solicitante
        solicitante = coluna_solicitante(descricao)
        if not solicitante:
            solicitante = "Victoria Utrini"

        rows.append({
            "Título": titulo_limpo,
            "Link do Card": link_card,
            "Link do Chamado": anex,
            "Data Done": data_done,
            "Protocolo": protocolo,
            "Tipo": tipo,
            "Solicitante": solicitante,
            "Pedido": descricao_limpa,
            "Observação": comentario,
            "Imagem Observação": imagem_comentario
        })

    # Cria ou usa modelo existente
    if os.path.exists(MODELO_PATH):
        wb = openpyxl.load_workbook(MODELO_PATH)
        ws = wb.active

        ws.delete_rows(2, ws.max_row)

        COLUNAS = [
            "Título", "Link do Card", "Link do Chamado", "Data Done", 
            "Protocolo", "Tipo", "Solicitante", "Pedido", 
            "Observação", "Imagem Observação"
        ]

        # Inserir os novos dados a partir da linha 2
        for i, r in enumerate(rows, start=2):
            for col_idx, col_name in enumerate(COLUNAS, 1):
                # Usamos o nome de coluna mapeado (Título, Tipo, etc.)
                ws.cell(row=i, column=col_idx, value=r.get(col_name, ""))

        ajustar_layout(ws)
        wb.save(ARQUIVO_EXCEL)

        
    else:
        # Se o MODELO não existe, criamos o DataFrame normalmente com cabeçalhos
        df = pd.DataFrame(rows)
        df.to_excel(ARQUIVO_EXCEL, index=False)
        wb = openpyxl.load_workbook(ARQUIVO_EXCEL)
        ws = wb.active
        ajustar_layout(ws)
        wb.save(ARQUIVO_EXCEL)

    print(f"\n✅ Arquivo salvo: {ARQUIVO_EXCEL}")

# ===================== EXECUÇÃO =====================
if __name__ == "__main__":
    gerar_planilha()
